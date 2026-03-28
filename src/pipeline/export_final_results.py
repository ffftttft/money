import argparse
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


ROOT = Path(__file__).resolve().parents[2]
FINAL_DIR = ROOT / "results" / "final_for_ppt"
SCORE_BASE = ROOT / "outputs" / "scores"
TEST_FRAME_DIR = ROOT / "data" / "frames" / "test"

SHORT = {
    "siglip": "SigLIP",
    "resnet101.a1h_in1k": "ResNet101",
    "vit_base_patch16_224.augreg2_in21k_ft_in1k": "ViT-B/16",
    "convnext_base.fb_in22k_ft_in1k": "ConvNeXt-B",
    "efficientnet_b0.ra_in1k": "EffNet-B0",
}

MODEL_ORDER = [
    "siglip",
    "resnet101.a1h_in1k",
    "vit_base_patch16_224.augreg2_in21k_ft_in1k",
    "convnext_base.fb_in22k_ft_in1k",
    "efficientnet_b0.ra_in1k",
]


def moving_average(x: np.ndarray, w: int = 9) -> np.ndarray:
    pad = w // 2
    x_pad = np.pad(x, (pad, pad), mode="edge")
    k = np.ones(w, dtype=np.float64) / float(w)
    return np.convolve(x_pad, k, mode="valid")


def safe(s: str) -> str:
    return s.replace("/", "_").replace(" ", "_")


def relative_to_root(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def load_method_scores(method: str):
    score_dir = SCORE_BASE / method
    raw = {}
    for p in sorted(score_dir.glob("*_test_scores.npy")):
        raw[p.name.replace("_test_scores.npy", "")] = np.load(p)
    names, scores = [], []
    for n in MODEL_ORDER:
        if n in raw:
            names.append(n)
            scores.append(raw[n])
    for n in sorted(raw.keys()):
        if n not in MODEL_ORDER:
            names.append(n)
            scores.append(raw[n])
    return names, scores


def save_summary_csv(method: str, names, scores):
    lines = ["model,samples,mean,std,p95,p99"]
    for n, s in zip(names, scores):
        lines.append(
            ",".join(
                [
                    SHORT.get(n, n),
                    str(len(s)),
                    f"{float(np.mean(s)):.6f}",
                    f"{float(np.std(s)):.6f}",
                    f"{float(np.percentile(s,95)):.6f}",
                    f"{float(np.percentile(s,99)):.6f}",
                ]
            )
        )
    (FINAL_DIR / f"{method}_03_model_summary_table.csv").write_text("\n".join(lines) + "\n", encoding="utf-8")
    save_summary_xlsx(method, names, scores)


def save_summary_xlsx(method: str, names, scores):
    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    headers = ["model", "samples", "mean", "std", "p95", "p99"]
    ws.append(headers)

    for n, s in zip(names, scores):
        ws.append(
            [
                SHORT.get(n, n),
                int(len(s)),
                float(np.mean(s)),
                float(np.std(s)),
                float(np.percentile(s, 95)),
                float(np.percentile(s, 99)),
            ]
        )

    # Header style
    hdr_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    for c in ws[1]:
        c.fill = hdr_fill
        c.font = Font(bold=True)

    # Highlight SigLIP row with a different color
    siglip_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "SigLIP":
            for col in range(1, 7):
                ws.cell(row=r, column=col).fill = siglip_fill
                ws.cell(row=r, column=col).font = Font(bold=True)
            break

    # Number formatting
    for r in range(2, ws.max_row + 1):
        for col in [3, 4, 5, 6]:
            ws.cell(row=r, column=col).number_format = "0.000000"

    # Column width
    widths = [16, 10, 12, 12, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    wb.save(FINAL_DIR / f"{method}_03_model_summary_table.xlsx")


def plot_tail(method: str, names, scores):
    p95 = np.array([np.percentile(s, 95) for s in scores], dtype=np.float64)
    p99 = np.array([np.percentile(s, 99) for s in scores], dtype=np.float64)
    labels = [SHORT.get(n, n) for n in names]
    x = np.arange(len(labels))
    w = 0.38

    fig, ax = plt.subplots(figsize=(12, 6))
    bars95 = ax.bar(x - w / 2, p95, width=w, label="P95", color="#5DAE8B")
    bars99 = ax.bar(x + w / 2, p99, width=w, label="P99", color="#6F8FB9")

    # Highlight SigLIP with distinct color from other four models.
    if "siglip" in names:
        i = names.index("siglip")
        bars95[i].set_facecolor("#E07A5F")
        bars99[i].set_facecolor("#C44536")
        bars95[i].set_edgecolor("black")
        bars99[i].set_edgecolor("black")
        bars95[i].set_linewidth(1.5)
        bars99[i].set_linewidth(1.5)
        ax.text(i, max(p99[i], p95[i]) + 1.0, "SigLIP", ha="center", va="bottom", fontsize=10, weight="bold")

    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=20, ha="right")
    ax.set_title(f"Tail Risk Comparison ({method.upper()})")
    ax.set_ylabel("Score")
    ax.legend()

    # Add numeric labels on bars (2 decimals).
    for rect in list(bars95) + list(bars99):
        h = rect.get_height()
        ax.text(
            rect.get_x() + rect.get_width() / 2.0,
            h + 0.01 * max(float(p99.max()), 1.0),
            f"{h:.2f}",
            ha="center",
            va="bottom",
            fontsize=9,
            rotation=0,
        )

    fig.tight_layout()
    fig.savefig(FINAL_DIR / f"{method}_02_tail_risk_comparison.png", dpi=300)
    plt.close(fig)


def export_siglip_detail(method: str, names, scores):
    if "siglip" not in names:
        return
    idx = names.index("siglip")
    s = scores[idx]
    
    # Scale score to 0-100 range using Option 1 (Global Baseline Non-linear Mapping)
    # We use a logistic curve mapping: f(x) = 100 / (1 + exp(-k * (x - x0)))
    # We calibrate it so the q95 threshold maps to ~30 (Moderate Risk).
    q95_raw = np.quantile(s, 0.95)
    median_raw = np.median(s)
    
    # Calibrate parameters
    # Let x0 be slightly above the median, and set k so that q95 maps to 30.
    # Risk = 100 * (1 / (1 + exp(-k * (s - x0))))
    x0 = median_raw + (q95_raw - median_raw) * 1.5 # Shift midpoint
    k = 1.8 / (q95_raw - median_raw + 1e-8)        # Slope factor
    
    s_scaled = 100 / (1 + np.exp(-k * (s - x0)))
    
    sm = moving_average(s_scaled, 9)
    th = float(np.quantile(sm, 0.95))
    binary = (sm >= th).astype(np.int32)

    # risk curve
    fig, ax = plt.subplots(figsize=(12, 4))
    x = np.arange(len(sm))
    ax.plot(x, sm, color="#2A9D8F", linewidth=1.6, label="Smoothed Risk Score (0-100)")
    ax.axhline(th, color="#D1495B", linestyle="--", linewidth=1.3, label="Q95 High Risk Threshold")
    ax.fill_between(x, sm, th, where=(sm >= th), color="#D1495B", alpha=0.2)
    ax.set_title(f"SigLIP Risk Dynamics ({method.upper()}) - Global Non-linear Mapping")
    ax.set_xlabel("Frame Index")
    ax.set_ylabel("Absolute Risk Level (0-100)")
    ax.set_ylim(0, 100) # Enforce visual axis scale to 0-100
    ax.legend(loc="upper right")
    fig.tight_layout()
    fig.savefig(FINAL_DIR / f"{method}_01_siglip_risk_curve.png", dpi=300)
    plt.close(fig)

    # status
    run = 0
    has_seg = False
    for v in binary:
        run = run + 1 if v == 1 else 0
        if run >= 5:
            has_seg = True
            break
    status = [
        f"method={method}",
        "model=siglip",
        f"samples={len(s_scaled)}",
        f"threshold_q95={th:.2f}",
        f"anomaly_ratio={float(binary.mean()):.6f}",
        f"is_anomaly={'yes' if has_seg else 'no'}",
    ]
    (FINAL_DIR / f"{method}_04_siglip_status.txt").write_text("\n".join(status) + "\n", encoding="utf-8")

    # topk and frames
    top = np.argsort(-s_scaled)[:10]
    lines = ["method,rank,frame_index,raw_score,risk_score"]
    rf_dir = FINAL_DIR / f"{method}_risk_frames"
    rf_dir.mkdir(parents=True, exist_ok=True)
    meta = ["rank,frame_index,raw_score,risk_score,test_time,video_time,src,dst"]
    for r, i in enumerate(top, 1):
        raw_sc = float(s[i])
        risk_sc = float(s_scaled[i])
        lines.append(f"{method},{r},{int(i)},{raw_sc:.6f},{risk_sc:.1f}")
        src = TEST_FRAME_DIR / f"frame_{int(i):06d}.jpg"
        test_sec = i / 2.0
        video_sec = 12 * 60 + test_sec
        tm = f"{int(video_sec//60):02d}:{int(video_sec%60):02d}"
        dst = rf_dir / f"rank{r}_idx{int(i):03d}_score{risk_sc:.1f}_t{tm}.jpg"
        if src.exists():
            dst.write_bytes(src.read_bytes())
        meta.append(
            f"{r},{int(i)},{raw_sc:.6f},{risk_sc:.1f},{test_sec:.3f},{video_sec:.3f},"
            f"{relative_to_root(src)},{relative_to_root(dst)}"
        )

    (FINAL_DIR / f"{method}_05_siglip_topk_frames.csv").write_text("\n".join(lines) + "\n", encoding="utf-8")
    (rf_dir / "meta.csv").write_text("\n".join(meta) + "\n", encoding="utf-8")


def cleanup_final_dir():
    FINAL_DIR.mkdir(parents=True, exist_ok=True)
    for p in FINAL_DIR.glob("*"):
        if p.is_file():
            p.unlink()
    # clean subdirs
    for p in FINAL_DIR.glob("*"):
        if p.is_dir():
            for f in p.rglob("*"):
                if f.is_file():
                    f.unlink()
            for d in sorted([d for d in p.rglob("*") if d.is_dir()], reverse=True):
                try:
                    d.rmdir()
                except OSError:
                    pass
            try:
                p.rmdir()
            except OSError:
                pass


def parse_args():
    p = argparse.ArgumentParser(description="Export final result pack from score folders")
    p.add_argument("--score-base", default=str(ROOT / "outputs" / "scores"), help="Base dir containing gaussian/knn")
    p.add_argument("--out-dir", default=str(ROOT / "results" / "final_for_ppt"), help="Output result folder")
    return p.parse_args()


def main():
    global FINAL_DIR, SCORE_BASE
    args = parse_args()
    FINAL_DIR = Path(args.out_dir)
    SCORE_BASE = Path(args.score_base)
    cleanup_final_dir()
    for method in ["gaussian", "knn"]:
        names, scores = load_method_scores(method)
        save_summary_csv(method, names, scores)
        plot_tail(method, names, scores)
        export_siglip_detail(method, names, scores)

    # progress snapshot
    progress = [
        "# 项目进度总览",
        "",
        "- 总步数：8",
        "- 已完成：8",
        "- 当前进度：100%",
        "",
        "已交付：Gaussian + KNN 两套最终图表与高风险帧。",
    ]
    (FINAL_DIR / "PROJECT_PROGRESS.md").write_text("\n".join(progress) + "\n", encoding="utf-8")
    print("saved:", FINAL_DIR)
    for p in sorted(FINAL_DIR.glob("*")):
        print("-", p.name)


if __name__ == "__main__":
    main()
